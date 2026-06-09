import urllib.request, json, datetime, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SURL='https://ocgzwbnuesjuagwbohtf.supabase.co/rest/v1'
SKEY='sb_publishable_H70ojBN_dEsaBzicie13Kg_2JzTDFJD'
H={'apikey':SKEY,'Authorization':'Bearer '+SKEY}

req=urllib.request.Request(SURL+'/operaciones?order=fecha_compra.asc',headers=H)
with urllib.request.urlopen(req) as r:
    ops=json.loads(r.read())

HEADERS=['Fecha compra','Plataforma compra','Modelo','Precio (£)','Envío (£)','Reparación (£)',
         'Forwarder (£)','Coste total (£)','Estado','Fecha publicación','Precio listado (£)',
         'Fecha venta','Plataforma venta','Ingreso neto (£)','Beneficio neto (£)','ROI (%)',
         'Cuenta','% Carlos','% Marta','% Dom']
DARK='1A1A1A'; LIGHT='F5F5F5'; GREEN='2ECC71'; RED='E74C3C'; BLUE='3498DB'

def hdr_cell(cell, text):
    cell.value=text
    cell.font=Font(bold=True,color='FFFFFF',name='Arial',size=9)
    cell.fill=PatternFill('solid',start_color=DARK)
    cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)

def make_sheet(wb, name, rows, is_first=False):
    ws=wb.active if is_first else wb.create_sheet(name)
    ws.title=name
    ws.freeze_panes='A2'
    ws.row_dimensions[1].height=30
    for ci,h in enumerate(HEADERS,1): hdr_cell(ws.cell(1,ci),h)
    for ri,o in enumerate(rows,2):
        estado=o.get('estado','')
        fill_color=LIGHT if ri%2==0 else 'FFFFFF'
        vals=[
            o.get('fecha_compra',''), o.get('plataforma_compra',''), o.get('modelo',''),
            o.get('precio_gbp') or '', o.get('gastos_envio_gbp') or '',
            o.get('reparacion_gbp') or '', o.get('gastos_forwarder_gbp') or '',
            o.get('coste_total_gbp') or '', estado,
            o.get('fecha_publicacion','') or '', o.get('precio_listado_gbp') or '',
            o.get('fecha_venta','') or '', o.get('plataforma_venta','') or '',
            o.get('ingreso_neto_gbp') if o.get('ingreso_neto_gbp') is not None else '',
            o.get('beneficio_neto_gbp') if o.get('beneficio_neto_gbp') is not None else '',
            round(o.get('roi_pct',0)*100,1) if o.get('roi_pct') is not None else '',
            o.get('cuenta',''),
            round((o.get('pct_carlos') or 0)*100,0),
            round((o.get('pct_marta') or 0)*100,0),
            round((o.get('pct_dom') or 0)*100,0),
        ]
        for ci,val in enumerate(vals,1):
            cell=ws.cell(ri,ci,val)
            cell.font=Font(name='Arial',size=9)
            cell.fill=PatternFill('solid',start_color=fill_color)
            if ci==9:
                col={'vendido':GREEN,'publicado':BLUE,'activo':'F39C12','devuelto':'E67E22'}.get(estado,'AAAAAA')
                cell.fill=PatternFill('solid',start_color=col)
                cell.font=Font(name='Arial',size=9,bold=True,color='FFFFFF')
            if ci==15 and val!='':
                try:
                    cell.font=Font(name='Arial',size=9,color=(GREEN if float(val)>=0 else RED),bold=True)
                except: pass
    widths=[12,14,38,9,9,9,9,11,10,13,11,12,14,11,11,7,8,7,7,6]
    for ci,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(ci)].width=w
    n=len(rows)+1
    ws.cell(n+2,3,'TOTAL VENDIDO (con precio)').font=Font(bold=True,name='Arial',size=9)
    ws.cell(n+2,14,f'=SUMIF(I2:I{n},"vendido",N2:N{n})').font=Font(bold=True,name='Arial',size=9,color=BLUE)
    ws.cell(n+2,15,f'=SUMIF(I2:I{n},"vendido",O2:O{n})').font=Font(bold=True,name='Arial',size=9,color=GREEN)
    ws.cell(n+3,3,'EN STOCK (coste)').font=Font(bold=True,name='Arial',size=9)
    ws.cell(n+3,8,f'=SUMIFS(H2:H{n},I2:I{n},"activo")+SUMIFS(H2:H{n},I2:I{n},"publicado")').font=Font(bold=True,name='Arial',size=9,color='E67E22')

wb=Workbook()
make_sheet(wb,'Todos',ops,is_first=True)
make_sheet(wb,'Carlos',[o for o in ops if o.get('cuenta')=='Carlos'])
make_sheet(wb,'Marta',[o for o in ops if o.get('cuenta')=='Marta'])

os.makedirs('reportes',exist_ok=True)
fecha=datetime.date.today().strftime('%Y-%m-%d')
out=f'reportes/Watches_Report_{fecha}.xlsx'
wb.save(out)
print(f'Guardado: {out} ({len(ops)} relojes)')
